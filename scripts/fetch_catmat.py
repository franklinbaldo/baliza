#!/usr/bin/env -S uv run --quiet
# /// script
# requires-python = ">=3.12"
# dependencies = [
#   "httpx>=0.27",
#   "openpyxl>=3.1",
# ]
# ///
"""Fetch CATMAT/CATSER catalog from gov.br and emit a typed JSON for the web UI.

Source (official, refreshed periodically by SERPRO/Compras):
  https://www.gov.br/compras/pt-br/acesso-a-informacao/consulta-detalhada/planilha-catmat-catser

CATMAT (~163K items) is deduplicated to PDM level (~10K kinds of materials)
because the per-SKU detail is irrelevant to a free-text resolver — the user
wants the catalog category, not a manufacturer-specific row.

CATSER (~3K items) is included in full.

Outputs (committed to git):
  web/src/lib/catmat-data.json   ~1.3 MB raw, ~300 KB gzipped
  web/src/lib/catmat-meta.json   { fetched_at, source_urls, counts, etag }

Refresh workflow:
  $ uv run scripts/fetch_catmat.py
  $ git add web/src/lib/catmat-data.json web/src/lib/catmat-meta.json
  $ git commit -m "chore(catmat): refresh from gov.br snapshot YYYY-MM-DD"

Detect upstream changes via the page RSS:
  https://www.gov.br/compras/pt-br/acesso-a-informacao/consulta-detalhada/planilha-catmat-catser/RSS
"""
from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

import httpx
import openpyxl

CATMAT_URL = "https://www.gov.br/compras/pt-br/acesso-a-informacao/consulta-detalhada/planilha-catmat-catser/catmat.xlsx"
CATSER_URL = "https://www.gov.br/compras/pt-br/acesso-a-informacao/consulta-detalhada/planilha-catmat-catser/catser.xlsx"

REPO_ROOT = Path(__file__).resolve().parent.parent
# Served as a static asset and fetched on demand from CatmatSearch — bundling it
# into the JS payload busted the 302 KB bundle-guard budget.
OUT_DATA = REPO_ROOT / "web" / "public" / "data" / "catmat.json"
OUT_META = REPO_ROOT / "web" / "public" / "data" / "catmat-meta.json"


def download(url: str) -> tuple[bytes, str | None]:
    print(f"  GET {url}")
    with httpx.Client(timeout=180.0, follow_redirects=True) as client:
        r = client.get(url)
        r.raise_for_status()
        return r.content, r.headers.get("etag")


def parse_catmat(blob: bytes) -> list[dict[str, str]]:
    """Parse CATMAT xlsx into PDM-deduplicated entries.

    Expected columns (from row 2, the header):
      0 codigo_grupo, 1 nome_grupo, 2 codigo_classe, 3 nome_classe,
      4 codigo_pdm, 5 nome_pdm, 6 codigo_item, 7 descricao_item, 8 codigo_ncm
    """
    wb = openpyxl.load_workbook(BytesIO(blob), read_only=True, data_only=True)
    ws = wb["Materiais"]
    rows = ws.iter_rows(values_only=True)
    next(rows)  # skip "Extração realizada em ..."
    next(rows)  # skip column headers

    seen: set[str] = set()
    entries: list[dict[str, str]] = []
    for r in rows:
        # Defensive bounds check: short rows on a future xlsx schema change
        # should be skipped, not crash the pipeline.
        if len(r) < 7 or r[6] is None:
            continue
        codigo_pdm = str(r[4]) if r[4] is not None else ""
        if not codigo_pdm or codigo_pdm in seen:
            continue
        seen.add(codigo_pdm)
        entries.append(
            {
                "code": codigo_pdm,
                "description": (r[5] or "").strip(),
                "type": "CATMAT",
            }
        )
    wb.close()
    return entries


def parse_catser(blob: bytes) -> list[dict[str, str]]:
    """Parse CATSER xlsx into entries (all rows; CATSER is small).

    Expected columns (from row 3, the header):
      0 tipo, 1 grupo_codigo, 2 grupo_nome, 3 classe_codigo, 4 classe_nome,
      5 codigo_servico, 6 descricao_servico, 7 situacao
    """
    wb = openpyxl.load_workbook(BytesIO(blob), read_only=True, data_only=True)
    ws = wb["Lista CATSER"]
    rows = ws.iter_rows(values_only=True)
    next(rows)  # "Lista CATSER"
    next(rows)  # "Extração realizada em ..."
    next(rows)  # column headers

    entries: list[dict[str, str]] = []
    for r in rows:
        if len(r) < 8 or r[5] is None:
            continue
        if (r[7] or "").strip().lower() != "ativo":
            continue
        entries.append(
            {
                "code": str(r[5]),
                "description": (r[6] or "").strip(),
                "type": "CATSER",
            }
        )
    wb.close()
    return entries


def main() -> int:
    print("Fetching CATMAT/CATSER from gov.br...")
    catmat_blob, catmat_etag = download(CATMAT_URL)
    catser_blob, catser_etag = download(CATSER_URL)

    print("Parsing CATMAT (PDM-dedup)...")
    catmat = parse_catmat(catmat_blob)
    print(f"  → {len(catmat)} unique PDMs")

    print("Parsing CATSER...")
    catser = parse_catser(catser_blob)
    print(f"  → {len(catser)} active services")

    entries = sorted(catmat + catser, key=lambda e: (e["type"], e["code"]))

    OUT_DATA.parent.mkdir(parents=True, exist_ok=True)
    OUT_DATA.write_text(
        json.dumps(entries, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    size_kb = OUT_DATA.stat().st_size / 1024
    print(f"Wrote {OUT_DATA.relative_to(REPO_ROOT)} ({size_kb:.0f} KB)")

    OUT_META.write_text(
        json.dumps(
            {
                "fetched_at": datetime.now(timezone.utc).isoformat(),
                "sources": {
                    "catmat": {"url": CATMAT_URL, "etag": catmat_etag},
                    "catser": {"url": CATSER_URL, "etag": catser_etag},
                },
                "counts": {"catmat": len(catmat), "catser": len(catser), "total": len(entries)},
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(f"Wrote {OUT_META.relative_to(REPO_ROOT)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
